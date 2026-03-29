"""
document_processors.py
-----------------------
Each class converts a specific file type to a dict of
{ page_number (int): PIL.Image } which is passed to the VLM extractor.

Page numbers are 1-indexed.
"""

from __future__ import annotations
import io
import os
import shutil
import subprocess
import tempfile
from abc import ABC, abstractmethod
from typing import Dict, List
from PIL import Image


# ─────────────────────────────────────────────────────────────────────────────
# Base
# ─────────────────────────────────────────────────────────────────────────────

class BaseDocumentProcessor(ABC):
    """
    Convert a file on disk to a dict of { page_number: PIL.Image }.
    Page numbers are 1-indexed integers.
    Single-page documents return {1: image}.
    """

    @abstractmethod
    def to_images(self, file_path: str) -> Dict[int, Image.Image]:
        ...

    # Convenience: return ordered list of images (used by legacy callers)
    def to_image_list(self, file_path: str) -> List[Image.Image]:
        page_dict = self.to_images(file_path)
        return [page_dict[k] for k in sorted(page_dict.keys())]


# ─────────────────────────────────────────────────────────────────────────────
# Image Processor  (PNG / JPG / BMP / TIFF / WEBP …)
# ─────────────────────────────────────────────────────────────────────────────

class ImageProcessor(BaseDocumentProcessor):
    """Single image file → {1: image}."""

    SUPPORTED_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".tif", ".webp"}

    def to_images(self, file_path: str) -> Dict[int, Image.Image]:
        img = Image.open(file_path).convert("RGB")
        return {1: img}


# ─────────────────────────────────────────────────────────────────────────────
# PDF Processor  — multipage aware
# ─────────────────────────────────────────────────────────────────────────────

class PDFProcessor(BaseDocumentProcessor):
    """
    Converts every page of a PDF to a PIL Image.
    Returns { page_number: PIL.Image } for ALL pages.
    Uses PyMuPDF (fitz) primary, pdf2image as fallback.
    """

    SUPPORTED_EXTENSIONS = {".pdf"}

    def to_images(self, file_path: str, dpi: int = 150) -> Dict[int, Image.Image]:
        try:
            import fitz  # PyMuPDF
            return self._with_fitz(file_path, dpi)
        except ImportError:
            pass

        try:
            from pdf2image import convert_from_path
            return self._with_pdf2image(file_path, dpi)
        except ImportError:
            raise RuntimeError(
                "PDF processing requires PyMuPDF (`pip install pymupdf`) "
                "or pdf2image (`pip install pdf2image`)."
            )

    def _with_fitz(self, file_path: str, dpi: int) -> Dict[int, Image.Image]:
        import fitz
        doc    = fitz.open(file_path)
        matrix = fitz.Matrix(dpi / 72, dpi / 72)
        pages  = {}
        for i, page in enumerate(doc, start=1):
            pix = page.get_pixmap(matrix=matrix, colorspace=fitz.csRGB)
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            pages[i] = img
        doc.close()
        print(f"[PDFProcessor] Loaded {len(pages)} page(s) from {os.path.basename(file_path)}")
        return pages

    def _with_pdf2image(self, file_path: str, dpi: int) -> Dict[int, Image.Image]:
        from pdf2image import convert_from_path
        images = convert_from_path(file_path, dpi=dpi)
        pages  = {i + 1: img.convert("RGB") for i, img in enumerate(images)}
        print(f"[PDFProcessor] Loaded {len(pages)} page(s) via pdf2image")
        return pages


# ─────────────────────────────────────────────────────────────────────────────
# Word Processor  (.docx / .doc)  — multipage aware
# ─────────────────────────────────────────────────────────────────────────────

class WordProcessor(BaseDocumentProcessor):
    """
    Converts DOCX/DOC → PDF via LibreOffice, then delegates to PDFProcessor
    so every page is captured individually.
    Falls back to rendering extracted text as paginated images.
    """

    SUPPORTED_EXTENSIONS = {".docx", ".doc"}

    def to_images(self, file_path: str) -> Dict[int, Image.Image]:
        # Primary: LibreOffice → PDF → PDFProcessor (preserves all pages)
        soffice = shutil.which("soffice") or shutil.which("libreoffice")
        if soffice:
            try:
                return self._via_libreoffice(file_path, soffice)
            except Exception as e:
                print(f"[WordProcessor] LibreOffice conversion failed: {e}, falling back.")

        # Fallback: python-docx text → paginated images
        return self._text_fallback(file_path)

    def _via_libreoffice(self, file_path: str, soffice: str) -> Dict[int, Image.Image]:
        with tempfile.TemporaryDirectory() as tmpdir:
            subprocess.run(
                [soffice, "--headless", "--convert-to", "pdf", "--outdir", tmpdir, file_path],
                check=True,
                capture_output=True,
            )
            pdf_files = [f for f in os.listdir(tmpdir) if f.endswith(".pdf")]
            if not pdf_files:
                raise RuntimeError("LibreOffice produced no PDF output.")
            pdf_path = os.path.join(tmpdir, pdf_files[0])
            pages = PDFProcessor().to_images(pdf_path)
            print(f"[WordProcessor] Converted to PDF → {len(pages)} page(s)")
            return pages

    def _text_fallback(self, file_path: str) -> Dict[int, Image.Image]:
        try:
            from docx import Document
        except ImportError:
            raise RuntimeError("Install python-docx: pip install python-docx")

        doc   = Document(file_path)
        lines = [p.text for p in doc.paragraphs if p.text.strip()]
        images = _text_lines_to_images(lines)
        pages  = {i + 1: img for i, img in enumerate(images)}
        print(f"[WordProcessor] Text fallback → {len(pages)} page(s)")
        return pages


# ─────────────────────────────────────────────────────────────────────────────
# Excel / CSV Processor
# ─────────────────────────────────────────────────────────────────────────────

class ExcelProcessor(BaseDocumentProcessor):
    """
    Converts each sheet to a text-table image.
    Returns { sheet_index: PIL.Image } (1-indexed).
    """

    SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv"}

    def to_images(self, file_path: str) -> Dict[int, Image.Image]:
        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".csv":
            import csv
            with open(file_path, newline="", encoding="utf-8", errors="replace") as f:
                rows = list(csv.reader(f))
            sheets = {"Sheet1": rows}
        else:
            try:
                import openpyxl
                wb     = openpyxl.load_workbook(file_path, data_only=True)
                sheets = {}
                for name in wb.sheetnames:
                    ws = wb[name]
                    sheets[name] = [[str(c.value or "") for c in row] for row in ws.iter_rows()]
            except ImportError:
                raise RuntimeError("Install openpyxl: pip install openpyxl")

        pages = {}
        idx   = 1
        for sheet_name, rows in sheets.items():
            lines      = [f"Sheet: {sheet_name}"] + ["  |  ".join(row) for row in rows]
            sheet_imgs = _text_lines_to_images(lines)
            for img in sheet_imgs:
                pages[idx] = img
                idx += 1

        print(f"[ExcelProcessor] {len(pages)} page(s) across {len(sheets)} sheet(s)")
        return pages


# ─────────────────────────────────────────────────────────────────────────────
# Generic Text Processor  (.txt)
# ─────────────────────────────────────────────────────────────────────────────

class GenericTextProcessor(BaseDocumentProcessor):
    """Reads plain text and renders as paginated PIL Images."""

    SUPPORTED_EXTENSIONS = {".txt"}

    def to_images(self, file_path: str) -> Dict[int, Image.Image]:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            lines  = f.read().splitlines()
        images = _text_lines_to_images(lines)
        pages  = {i + 1: img for i, img in enumerate(images)}
        print(f"[GenericTextProcessor] {len(pages)} page(s)")
        return pages


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _text_lines_to_images(
    lines: list,
    font_size: int = 16,
    margin: int = 40,
    max_lines_per_page: int = 60,
) -> List[Image.Image]:
    """Render a list of text strings into one or more A4-like white PIL Images."""
    from PIL import ImageDraw, ImageFont

    try:
        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", font_size)
    except Exception:
        font = ImageFont.load_default()

    chunks = [lines[i: i + max_lines_per_page] for i in range(0, max(len(lines), 1), max_lines_per_page)]
    images = []
    width, height = 1240, 1754  # A4 at 150 dpi

    for chunk in chunks:
        img  = Image.new("RGB", (width, height), color=(255, 255, 255))
        draw = ImageDraw.Draw(img)
        y    = margin
        for line in chunk:
            draw.text((margin, y), line, fill=(30, 30, 30), font=font)
            y += font_size + 4
        images.append(img)

    return images


def get_processor_for_file(extension: str) -> BaseDocumentProcessor:
    """Return the appropriate processor for the given file extension."""
    ext = extension.lower()
    if ext in ImageProcessor.SUPPORTED_EXTENSIONS:
        return ImageProcessor()
    if ext in PDFProcessor.SUPPORTED_EXTENSIONS:
        return PDFProcessor()
    if ext in WordProcessor.SUPPORTED_EXTENSIONS:
        return WordProcessor()
    if ext in ExcelProcessor.SUPPORTED_EXTENSIONS:
        return ExcelProcessor()
    if ext in GenericTextProcessor.SUPPORTED_EXTENSIONS:
        return GenericTextProcessor()
    return ImageProcessor()